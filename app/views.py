from django.shortcuts import render
from .models import StationDetails,Authentication,Station_store,Semlog_data
import os
from django.core.exceptions import ValidationError
import requests
import openpyxl
from django import forms
import os
from django.contrib.auth.decorators import login_required
# from datetime import date
import datetime
from datetime import timedelta
from openpyxl.utils import get_column_letter
import pdb,openpyxl
import pandas as pd
import xlwt
import xlsxwriter
from django.shortcuts import render_to_response
from django.http import HttpResponse,HttpResponseNotFound,Http404
from django import forms
#from .forms import NameForm
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, get_object_or_404, redirect
from django.db import connection
from django.core.files.storage import FileSystemStorage,default_storage
from django.contrib import messages
from tdrift.settings import EMAIL_HOST_USER
from . import forms
from django.core.mail import send_mail


sr1_list=['CPETA','CUDDAPAH','CUDDAPAH 765','GAZUWAKA HVDC','GAZUWAKA','GHANAPUR','GOOTY','KHAMMAM','KURNOOL 765kV','MAHESHWARAM','NELLORE','Nellore PS','NIZAMABAD','NP KUNTA (PG)','NSAGAR','SRIKAKULAM','TALCHER HVDC','VIJAYAWADA','WARANGAL']

sr2_list=['ARASUR','BIDADI','DHARMAPURI','HASAN','HIRIYUR','HOSUR','KALAVINTHAPATTU','KARAIKUDI','KOCHI','KOLAR','KOLAR HVDC','KOZHIKODE','KUDGI (PG)','MADHUGIRI','MADURAI','MYSORE','MUNIRABAD','NAGAPATTINAM','NARENDRA','PALAKKAD','PAVAGADA','PUDUCHERRY','PUGALUR','RAICHUR 765kV','SALEM','SOMANAHALLI','SRIPERUMBUDUR','THIRUVALAM','Thiruvalem 765','TIRUNELVELI','TRICHUR','TRICHY','TRIVANDRUM','TUTICORIN GIS','TUTICORIN PS','UDUMALPET','YELAHANKA']
solar=['ADANI','AVAADA','KREDL']

station_names={"APcarbides":"AP CARBIDES","ALAMPUR":"ALAMPUR X Roads","BELLARY_PS":"BELLARY PS","COASTAL":"COASTAL ENERGEN","CUDDAPAH765":"CUDDAPAH 765","GAZUWAKA_HVDC":"GAZUWAKA HVDC","HOSUR_230":"HOSUR230","ILFS":"IL&FS","KAIGA_U12":"KAIGA U12","KAIGA_U34":"KAIGA U34","KCP_CEMENTS":"KCP CEMENTS","KOLAR_HVDC":"KOLAR HVDC","KUDGI_PG":"KUDGI(PG)","KUDGI_STPP":"KUDGI STPP (NTPC)","KURNOOL_765":"KURNOOL 765kV","KUDANKULAM_NPP":"KUDANKULAM NPP","LOWERSILERU":"LOWER SILERU","NSAGAR_RC":"N'SAGAR RC","NELLORE_MANABOLU":"NELLORE(MANABOLU)","NELLORE_PS":"Nellore PS","NLCTS1_EXP":"NLC TS-I EXP","NLCTS2_EXP":"NLC TS-II EXP","NLCTS_2STG1":"NLC TS-II STG1","NLCTS_2STG2":"NLC TS-II STG2","NLC_TN":"NLC(TN)","NPKUNTA":"NP KUNTA (PG)","NSAGAR_220":"NSAGAR -TS (220kV)","ORANGE":"ORANGE SIRONJ","VALLUR":"NTECL VALLUR","PVG_KSPDCL":"PAVAGADA KSPDCL","PUGALUR_HVDC":"PUGALUR HVDC","RAICHUR_765":"RAICHUR 765kV","RSTPS_STG12":"RSTPS STG 12","RSTPS_STG3":"RSTPS STG 3","SGPL":"SGPL (NCC PPL)","SP_KOIL":"SP KOIL","SRISAILAM_RB":"SRISAILAM RB","SRISAILAM_LB":"SRISAILEM LB","SURYAPETA_220":"SURYAPETA(220kV)","SURYAPETA_400":"SURYAPETA(400kV)","SVCHATRAM":"SV CHATRAM","TALCHER_HVDC":"TALCHER HVDC","TELANGANA_STPP":"TELANGANA STPP","THERMAL_POWERTECH":"THERMAL POWERTECH","Thiruvalem_765":"Thiruvalem 765","TUTICORIN_GIS":"TUTICORIN GIS","TUTICORIN_PS":"TUTICORIN PS","UPPER_SILERU":"UPPER SILERU","VEMAGIRI_PG":"VEMAGIRI(PG)","PARAMPUJYA":"ADANI"}

pvg_solar={"KSPDCL":"KSPDCL","FORTUMFINN":"FINN","ACME_REWA":"REWA","ACME_KURU":"KURU","SOFT_BANK":"SOFT BANK","PARAMPUJYA":"ADANI"}

pvg_solarlist=["KSPDCL","AZURE","FINN","YARROW","FORTUM","AVAADA","RENEW","ADYAH","KREDL","TATA","SOFT BANK","ADANI","KURU","REWA"]
result={}

# def subscribe(request):
#     if request.method == 'POST':
#         sub=request.POST.get('email')
#         subject = 'Welcome to DataFlair'
#         message = 'Hope you are enjoying your Django Tutorials'
#         recepient = str(sub)

#         send_mail(subject, 
#             message, EMAIL_HOST_USER, [recepient], fail_silently = False)
        
#         return render(request,'success.html')
#     # return render(request, 'subscribe/index.html', {'form':sub})
@login_required()
def upload(request):
    permitted_user=request.session['admin']
    if permitted_user =="SRLDC_MO":
        

            # context['url']= fs.url(filename)
        return render(request,"upload.html")
    else:
        text=""" <h1 style="color:green";> Permission denied by admin </h1> """
        return HttpResponse(text)
@login_required()
def uploaded_file(request):
    context={}
    nofileup=[]
    unsupformat=[]
    invalid=[]
    user=request.user
    uploaded_file = request.FILES.get('document',None)
    # If no file uploads this if condition
    if not uploaded_file:
        nofileup.append("nofile")
        return render(request,'upload.html',{"nofileup":nofileup})
    #File uploaded but invalid format
    ext = os.path.splitext(uploaded_file.name)[1]  # [0] returns path+filename
    valid_extensions = ['.xlsx', '.xls']
    if not ext.lower() in valid_extensions:
        unsupformat.append(ext)
        return render(request,'upload.html',{"unsupformat":unsupformat})
    
    df=pd.read_excel(uploaded_file,skiprows=3)
    #Correct format but different file uploaded
    try:
        length=len(df['Station Name'])
        if length > 1500:
            
            for i in range(0,length):

                stat_data=Semlog_data(
                    Utility_Name=df['Utility Name'][i],
                    Station_Name=df['Station Name'][i],
                    Location= df['Location'][i],
                    Description=df['Description'][i],
                    Meter_No=df['Meter No'][i]
                    )
                stat_data.save()
            for row in Semlog_data.objects.all():
                if Semlog_data.objects.filter(Meter_No=row.Meter_No,        Station_Name=row.Station_Name).count() > 1:
                    row.delete()
            return render(request,'Success.html')
    except Exception as e:
        invalid.append(e)
        return render(request,"upload.html",{"invalid":invalid})

   
    
@login_required()
def user_fileupload(request):
    if request.method == 'POST':
        context={}
        nofileup=[]
        unsupformat=[]
        invalid=[]
        user=request.user
        uploaded_file = request.FILES.get('document',None)
        #No file uploaded
        if not uploaded_file:
            nofileup.append("nofile")
            return render(request,'excel_upload.html',{"nofileup":nofileup})
        #File uploaded but invalid format
        ext = os.path.splitext(uploaded_file.name)[1]  # [0] returns path+filename
        valid_extensions = ['.xlsx', '.xls']
        if not ext.lower() in valid_extensions:
            unsupformat.append(ext)
            return render(request,'excel_upload.html',{"unsupformat":unsupformat})
        try:
            df=pd.read_excel(uploaded_file)
            length=len(df['Station_name'])
            if length >= 1:
                startTime = datetime.timedelta(0, 54915)
                startTime = (datetime.datetime.min + startTime).time()
                file_check1=[]
                file_check2=[]
                file_check3=[]
                # for string type refernce in line no 159
                correction='YES' 
                # New Correc done here
                
                date_modified=[]
               
                for i in range(0,len(df['Station_name'])):
                    if (len(str(df['Date_of_Check(dd/mm/yy)'][i]))!=0):

                        
                        continue
                        
                    # if (df['Date_of_Check(dd/mm/yy)'][i])== type(pd.Timestamp('2017-01-01T12')):
                    else:
                        
                        file_check1.append(df['Station_name'][i])
                        return render(request,"excel_upload.html",{"file_check1":file_check1})
                    # if df['Date_of_Check(dd/mm/yy)'][i] is not None:
                    #     continue
                    # else:
                    #     file_check1.append(df['Station_name'][i])
                    #     return render(request,"excel_upload.html",{"file_check1":file_check1})
            
                for i in range(0,len(df['Station_name'])):
                    if type(df['GPS(hh:mm:ss)'][i])== type(startTime)  :
                        continue
                    else:
                        file_check2.append(df['Station_name'][i])
                        return render(request,"excel_upload.html",{"file_check2":file_check2})
                for i in range(0,len(df['Station_name'])):
                    if type(df['Meter_time(hh:mm:ss)'][i])== type(startTime) :
                        continue
                    else:
                        file_check2.append(df['Station_name'][i])
                        return render(request,"excel_upload.html",{"file_check2":file_check2})

                for i in range(0,len(df['Station_name'])):
                    if type(df['Correction_Needed(Yes/No)'][i])== type(correction) :
                        correction=df['Correction_Needed(Yes/No)'][i].upper()
                        if correction == 'YES' or 'NO' :
                            continue
                        else:
                            file_check3.append(correction)
                            return render(request,"excel_upload.html",{"file_check3":file_check3})
                    else:
                        file_check3.append(df['Station_name'][i])
                        return render(request,"excel_upload.html",{"file_check3":file_check3})
        except Exception as e:
            invalid.append(e)
            return render(request,"excel_upload.html",{"invalid":invalid})
        #Left values in excel
         
        na_free=df.dropna(subset=['GPS(hh:mm:ss)','Meter_time(hh:mm:ss)','Date_of_Check(dd/mm/yy)','Correction_Needed(Yes/No)'])
        
        utility_name=Semlog_data.objects.filter(Station_Name__in=[na_free['Station_name'][0]]).values_list('Utility_Name','Description')
        
        only_na = df[~df.index.isin(na_free.index)]
        #na means nullable values
        new_dates=[]
        
        for i in range(length):
            # new_dates.append(df['Date_of_Check(dd/mm/yy)'][i])
            if type(df['Date_of_Check(dd/mm/yy)'][i])== type(pd.Timestamp('2017-01-01T12')):
               
                changed_type=df['Date_of_Check(dd/mm/yy)'][i].strftime('%d/%m/%Y')
                new_dates.append(datetime.datetime.strptime(changed_type,'%d/%m/%Y').date())
               
            else:
                new_dates.append(datetime.datetime.strptime(df['Date_of_Check(dd/mm/yy)'][i],'%d/%m/%Y').date())
        
       
        row_list =[]  
        for index, rows in only_na.iterrows():
            my_list =[rows.Station_name, rows.Meter_no]
            row_list.append(my_list) 
        na_free.index = [i for i in range(0,length)]
        length=len(na_free['Station_name'])
        
        for i in range(length):
            
            gps=datetime.datetime.strptime(str(na_free['GPS(hh:mm:ss)'][i]),'%H:%M:%S')
            meterdri=datetime.datetime.strptime(str(na_free['Meter_time(hh:mm:ss)'][i]),'%H:%M:%S')
            
            if gps >= meterdri:
                diff=gps-meterdri
                
                new_drift=datetime.datetime.strptime(str(diff),'%H:%M:%S').time()
            else:
                diff=meterdri-gps
                new_drift=datetime.datetime.strptime(str(diff),'%H:%M:%S').time()
                
                
            if na_free['GPS(hh:mm:ss)'][i] > na_free['Meter_time(hh:mm:ss)'][i]:
                meter_stat='SLOW'
            else:
                meter_stat='FAST'
           
            upload_data=StationDetails(
                utility_name=utility_name[i][0],
                station_name=na_free['Station_name'][i],
                description=utility_name[i][1],
                Meter_no=na_free['Meter_no'][i],
                gps=na_free['GPS(hh:mm:ss)'][i],
                meter_drift=na_free['Meter_time(hh:mm:ss)'][i],
                meter_difference=new_drift,
                meter_status=meter_stat,
                dateofchecking=new_dates[i],
                # na_free['Date_of_Check(dd/mm/yy)'][i],
                dateofupload=datetime.datetime.now(),
                correction_needed=na_free['Correction_Needed(Yes/No)'][i],
                remarks=na_free['Remarks'][i]
                )
            upload_data.save()
        if  only_na.empty == False:
            return render(request,"dropped_values.html",{"dropped":row_list})
        else:
            return render(request,"success.html")

    else:
        return render(Http404)
@login_required()
def home_page(request):
    try:
        if request.user.username =='SRLDC_MO':
            request.session['admin']= request.user.username
            updated_stations =Semlog_data.objects.filter().order_by('Station_Name').values_list('Station_Name').distinct()
            updated_list=[]
            for i in range(len(updated_stations)):
                updated_list.append(updated_stations[i][0])

            return render(request, 'index.html',{'updated_details': updated_list})
       
        elif request.user.username=='S_SR1PG':
            request.session['admin']= request.user.username
            return render(request, 'index1.html')

        elif request.user.username=='S_SR2PG':
            request.session['admin']= request.user.username
            return render(request, 'index1.html')

        else:
            request.session['admin']= request.user.username
            
            station = str(request.user.username)
            request.session['station1']=station
            global station_names,pvg_solar,pvg_solarlist

            if station not in station_names:
                if station not in pvg_solar:
                    station=station
                else:
                    station=pvg_solar[station]
            else:
                station=station_names[station]

            
           
            if station =="KSPDCL":
                new=Semlog_data.objects.filter(Description__icontains=station).values_list('Station_Name','Description','Meter_No')
                new=new.exclude(Description__icontains="MAIN").values_list('Station_Name','Description','Meter_No')
                new=new.exclude(Description__icontains="CHECK").values_list('Station_Name','Description','Meter_No')
            elif station =="PG_SOLAR":
                station="KSPDCL"
                new=Semlog_data.objects.filter(Description__icontains=station).values_list('Station_Name','Description','Meter_No')
                new=new.exclude(Description__icontains="STANDBY").values_list('Station_Name','Description','Meter_No')
            elif station =="FORTUM":
                new=Semlog_data.objects.filter(Description__icontains=station).values_list('Station_Name','Description','Meter_No')
                new=new.exclude(Description__icontains="FINN").values_list('Station_Name','Description','Meter_No')   
            elif station=="AZURE":
                new=Semlog_data.objects.filter(Description__icontains=station).values_list('Station_Name','Description','Meter_No')
                new=new.exclude(Description__icontains="ANANTHAPURAM").values_list('Station_Name','Description','Meter_No')
            elif station=="TATA":
                new=Semlog_data.objects.filter(Description__icontains=station).values_list('Station_Name','Description','Meter_No')
                new=new.exclude(Description__icontains="ANANTHAPURAM").values_list('Station_Name','Description','Meter_No')

            elif station in pvg_solarlist:
                new=Semlog_data.objects.filter(Description__icontains=station).values_list('Station_Name','Description','Meter_No')
            
            
            

            
            elif station=="ACMEKRNL":
                new=Semlog_data.objects.filter(Description__icontains=station).values_list('Station_Name','Description','Meter_No')
            elif station=="ACMEHIS":
                new=Semlog_data.objects.filter(Description__icontains=station).values_list('Station_Name','Description','Meter_No')
            elif station=="ACMEBHI":
                new=Semlog_data.objects.filter(Description__icontains=station).values_list('Station_Name','Description','Meter_No')
            elif station=="FRV":
                new=Semlog_data.objects.filter(Description__icontains=station).values_list('Station_Name','Description','Meter_No')
                new=new.exclude(Description__icontains="FRV8").values_list('Station_Name','Description','Meter_No')
            elif station=="FRV8":
                new=Semlog_data.objects.filter(Description__icontains=station).values_list('Station_Name','Description','Meter_No')
            elif station=="TATA_ANTP":
                station="TATA"
                new=Semlog_data.objects.filter(Description__icontains=station).values_list('Station_Name','Description','Meter_No')
                new=new.exclude(Description__icontains="PAVAGADA").values_list('Station_Name','Description','Meter_No')
            elif station=="AZURE_ANTP":
                station="AZURE"
                new=Semlog_data.objects.filter(Description__icontains=station).values_list('Station_Name','Description','Meter_No')
                new=new.exclude(Description__icontains="PAVAGADA").values_list('Station_Name','Description','Meter_No')  

            else:
                new=Semlog_data.objects.filter(Station_Name__in=[station]).values_list('Station_Name','Description','Meter_No')
            
            new1=list(new)
            for i in range(0,len(new1)):
                stat=list(new1[i])
                stat.insert(0,i+1)
                new1[i]=tuple(stat)
            return render_to_response('hour.html',{"list": new1})
    except:
        raise Http404("Page does not exist")

@login_required()
def stationdata(request1):
    try:

        if request1.method =='POST':

            Gps=request1.POST.getlist('GPS')
            meter_dri=request1.POST.getlist('METER DRIFT')
            # difference=request1.POST.getlist('difference'),
            date=request1.POST.getlist('DOC')
            
            correction=request1.POST.getlist('station')
            remarks=request1.POST.getlist('REMARKS')
            station=request1.POST.getlist('station name')
            global station_names,pvg_solar,pvg_solarlist
            old_station_name=str(request1.user.username)

            if old_station_name not in station_names:
                if old_station_name not in pvg_solar:

                    old_station_name=old_station_name
                else: 
                    old_station_name=pvg_solar[old_station_name]
            else:
                old_station_name=station_names[old_station_name]

            if old_station_name =="KSPDCL":
                    utility_name=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Utility_Name','Description')
                    utility_name=utility_name.exclude(Description__icontains="MAIN").values_list('Utility_Name','Description')
                    utility_name=utility_name.exclude(Description__icontains="CHECK").values_list('Utility_Name','Description')

            elif old_station_name =="PG_SOLAR":
                old_station_name="KSPDCL"
                utility_name=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Utility_Name','Description')
                utility_name=utility_name.exclude(Description__icontains="STANDBY").values_list('Utility_Name','Description')
            elif old_station_name =="TATA":
                utility_name=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Utility_Name','Description')
                utility_name=utility_name.exclude(Description__icontains="ANANTHAPURAM").values_list('Utility_Name','Description')
            elif old_station_name =="AZURE" :
                utility_name=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Utility_Name','Description')
                utility_name=utility_name.exclude(Description__icontains="ANANTHAPURAM").values_list('Utility_Name','Description')
            elif old_station_name =="FORTUM":
                utility_name=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Utility_Name','Description')
                utility_name=utility_name.exclude(Description__icontains="FINN").values_list('Utility_Name','Description')

            elif old_station_name in pvg_solarlist:
                utility_name=list(Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Utility_Name','Description'))

            else:
                utility_name=list(Semlog_data.objects.filter(Station_Name__in=[station[1]]).values_list('Utility_Name','Description')) 
                
            j=1
              
            for i in range(0,len(Gps)):
                new_gps=datetime.datetime.strptime(Gps[i],'%H:%M:%S').time()
                new_meterdri=datetime.datetime.strptime(meter_dri[i],'%H:%M:%S').time()
                gps=datetime.datetime.strptime(Gps[i],'%H:%M:%S')
                meterdri=datetime.datetime.strptime(meter_dri[i],'%H:%M:%S')
                if gps >= meterdri:
                    diff=gps-meterdri
                
                    new_drift=datetime.datetime.strptime(str(diff),'%H:%M:%S').time()
                else:
                    diff=meterdri-gps
                    new_drift=datetime.datetime.strptime(str(diff),'%H:%M:%S').time()
                
                
                if new_gps > new_meterdri:
                    meter_stat='SLOW'
                else:
                    meter_stat='FAST' 
                stat_data=StationDetails(
                    utility_name=utility_name[i][0],
                    station_name=station[j],
                    description=utility_name[i][1],
                    Meter_no=station[j+2],
                    gps=datetime.datetime.strptime(Gps[i],'%H:%M:%S').time(),
                    meter_drift=datetime.datetime.strptime(meter_dri[i],'%H:%M:%S').time(),
                    meter_difference=new_drift,
                    meter_status=meter_stat,
                    dateofchecking=datetime.datetime.strptime(date[i],'%m/%d/%Y').date(),
                    dateofupload=datetime.datetime.now().date(),
                    correction_needed=correction[i],
                    remarks=remarks[i]
                )
                
                stat_data.save()
                j=j+4
            return render(request1, 'success.html')
            
        else:
            #pdb.set_trace()
            return HttpResponseNotFound(request1, 'error_404.html')
    except:
        raise HttpResponseNotFound(request1, 'error_404.html')



@login_required()    
def formsubmit(request5):
    
    #pdb.set_trace()
    try:

        station = request5.GET.get('station')
        request5.session['station1']=station
    
        new=Semlog_data.objects.filter(Station_Name__in=[station]).values_list('Station_Name','Description','Meter_No')
        
        new1=list(new)

        for i in range(0,len(new1)):
            stat=list(new1[i])
            stat.insert(0,i+1)
            new1[i]=tuple(stat)  
        return render_to_response('hour.html',{"list": new1})
    except:
        raise Http404("Page not exist")
@login_required()
def station_finaldata(request1):
    try:

        filename="Semlog.xlsx"
        df=pd.read_excel(filename,skiprows=3)
        stat=list(df['Location'])
        station_name=request1.session['station'][0][0]
        
        
        for i in range(len(df['Station Name'])):
            if station_name==stat[i]:
                station_value= df['Station Name'][i]
            else:
                continue 
        posts =StationDetails.objects.filter().order_by('-meter_drift').all()
        posts=posts.extra(where=['station_name=%s'], params=[station_value])
        return render(request1, 'fulldetails.html', {'details': posts})
    except:
        raise Http404("Page does not exist")
@login_required()
def meter_details(request1):
    try:
        # station_name=request1.session['station1']
        station_name=request1.user.username
        global sr1_list,sr2_list
        
        if request1.user.username =="S_SR1PG":

            sr1_details=StationDetails.objects.filter(station_name__in=[i for i in sr1_list]).values_list('station_name','Meter_no','gps','meter_drift','dateofchecking','correction_needed','remarks').values()
            sr1_details =sr1_details.filter().order_by('-meter_drift').values_list('station_name','Meter_no','gps','meter_drift','dateofchecking','correction_needed','remarks').values()
            # posts=posts.extra(where=['station_name=%s'], params=[station_name])
            

            return render(request1, 'fulldetails.html', {'details': sr1_details})
        
        elif request1.user.username =="S_SR2PG":
            
            sr2_details=StationDetails.objects.filter(station_name__in=[i for i in sr2_list]).values_list('station_name','Meter_no','gps','meter_drift','dateofchecking','correction_needed','remarks').values()
            sr2_details =sr2_details.filter().order_by('-meter_drift').values_list('station_name','Meter_no','gps','meter_drift','dateofchecking','correction_needed','remarks').values()
            # posts=posts.extra(where=['station_name=%s'], params=[station_name])
            return render(request1, 'fulldetails.html', {'details': sr2_details})
        
        elif request1.user.username =="SRLDC_MO":
            rldc_details=StationDetails.objects.values_list('station_name','Meter_no','gps','meter_drift','dateofchecking','remarks').values()
            rldc_details= rldc_details.filter().order_by('-meter_drift').values_list('station_name','Meter_no','gps','meter_drift','dateofchecking','correction_needed','remarks').values()
            return render(request1, 'fulldetails.html', {'details':rldc_details})
        else:
            stat_details=StationDetails.objects.filter(station_name__in=[station_name]).values_list('station_name','Meter_no','gps','meter_drift','dateofchecking','correction_needed','remarks').values()
            stat_details=stat_details.filter().order_by('-meter_drift').values_list('station_name','Meter_no','gps','meter_drift','dateofchecking','correction_needed','remarks').values()
            return render(request1, 'fulldetails.html', {'details':stat_details})
    except:
        raise Http404("Page does not exist")
@login_required()                
def month_detailspage(request3):
    try:
        return render(request3,'meter_details.html')
    except:
        raise Http404("Page does not exist")

@login_required()
def date_wise(request):
    try:
        
        return render(request,'date_range.html')
    except:
        raise Http404("Page does not exist")   
@login_required()
def month_wise(request5):
    try:
        month1=request5.GET.get('month')
        first_day=[]
        last_day=[]
        threshold=request5.GET.get('threshold')
        
        new_threshold=datetime.datetime.strptime(threshold,'%H:%M:%S').time()
        
        year=int(month1[3:])
        global sr1_list,sr2_list,pvg_solar,pvg_solarlist

        request5.session["selected_month"]=month1
        def last_day_of_month(any_day):
            next_month = any_day.replace(day=28) + datetime.timedelta(days=4)
            return next_month - datetime.timedelta(days=next_month.day)
        
        for month in range(1, 13):
            todayDate = datetime.date(year,month,1)
            if todayDate.day > 25:
                todayDate += datetime.timedelta(7)
            first_day.append(todayDate.replace(day=1))
            last_day.append(last_day_of_month(datetime.date(year, month, 1)))
        if month1[0:2]=='01':
            start_date=first_day[0]
            end_date=last_day[0]
        elif month1[0:2]=='02':
            start_date=first_day[1]
            end_date=last_day[1]
        elif month1[0:2]=='03':
            start_date=first_day[2]
            end_date=last_day[2]
        elif month1[0:2]=='04':
            start_date=first_day[3]
            end_date=last_day[3]
        elif month1[0:2]=='05':
            start_date=first_day[4]
            end_date=last_day[4]
        elif month1[0:2]=='06':
            start_date=first_day[5]
            end_date=last_day[5]
        elif month1[0:2]=='07':
            start_date=first_day[6]
            end_date=last_day[6]
        elif month1[0:2]=='08':
            start_date=first_day[7]
            end_date=last_day[7]
        elif month1[0:2]=='09':
            start_date=first_day[8]
            end_date=last_day[8]
        elif month1[0:2]=='10':
            start_date=first_day[9]
            end_date=last_day[9]
        elif month1[0:2]=='11':
            start_date=first_day[10]
            end_date=last_day[10]
        else:
            start_date=first_day[11]
            end_date=last_day[11]

        old_station_name=request5.user.username

        if request5.user.username == 'SRLDC_MO':

            total_details=list(StationDetails.objects.filter(dateofchecking__range=(start_date, end_date),meter_difference__range=[new_threshold,'23:59:59']).values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference','dateofchecking','correction_needed','meter_status').order_by('-meter_difference').distinct())
            
            
            # stat_details=stat_details.extra(where=['start_date=%s','end_date=%s'], params=[start_date,end_date])
            # total_details=list(total_details.filter().order_by('-dateofchecking').values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference','dateofchecking','correction_needed','meter_status'))
            location_set=[]
            for i in range(0,len(total_details)):
                location=[]
                location=Semlog_data.objects.filter(Description__icontains=total_details[i][2]).values_list('Location','Description')
                   
                location_set.append(location[0])

            for i in range(0,len(total_details)):
                stat=list(total_details[i])
                stat[4]=str(stat[4])
                stat[5]=str(stat[5])
                stat[6]=str(stat[6])
                stat.insert(4,str(location_set[i][0]))     
                total_details[i]=tuple(stat)

            return render(request5, 'meter_details.html', {'totdetails': total_details})

        elif request5.user.username == 'S_SR1PG':
            total_details=list(StationDetails.objects.filter(station_name__in=[i for i in sr1_list],dateofchecking__range=(start_date, end_date),meter_difference__range=[new_threshold,'23:59:59']).values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference','dateofchecking','correction_needed','meter_status').order_by('-meter_difference').distinct())

            location_set=[]
            for i in range(0,len(total_details)):
                location=[]
                location=Semlog_data.objects.filter(Description__icontains=total_details[i][2]).values_list('Location','Description')
                   
                location_set.append(location[0])


            for i in range(0,len(total_details)):
                stat=list(total_details[i])
                stat[4]=str(stat[4])
                stat[5]=str(stat[5])
                stat[6]=str(stat[6])
                stat.insert(4,str(location_set[i][0]))     
                total_details[i]=tuple(stat)

            return render(request5, 'meter_details.html', {'totdetails': total_details})

        elif request5.user.username == 'S_SR2PG':
            total_details=list(StationDetails.objects.filter(station_name__in=[i for i in sr2_list],dateofchecking__range=(start_date, end_date),meter_difference__range=[new_threshold,'23:59:59']).values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference','dateofchecking','correction_needed','meter_status').order_by('-meter_difference').distinct())

            location_set=[]
            for i in range(0,len(total_details)):
                location=[]
                location=Semlog_data.objects.filter(Description__icontains=total_details[i][2]).values_list('Location','Description')
                   
                location_set.append(location[0])


            for i in range(0,len(total_details)):
                stat=list(total_details[i])
                stat[4]=str(stat[4])
                stat[5]=str(stat[5])
                stat[6]=str(stat[6])
                stat.insert(4,str(location_set[i][0]))     
                total_details[i]=tuple(stat)

            return render(request5, 'meter_details.html', {'totdetails': total_details})
        else:
            if old_station_name not in station_names:
                if old_station_name not in pvg_solar:
                    old_station_name=old_station_name
                else:
                    old_station_name=pvg_solar[old_station_name]
            else:
                old_station_name=station_names[old_station_name]

            if old_station_name =="KSPDCL":
                meter_list=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Meter_No')
                meter_list=meter_list.exclude(Description__icontains="MAIN").values_list('Meter_No')
                meter_list=meter_list.exclude(Description__icontains="CHECK").values_list('Meter_No')
            elif old_station_name =="PG_SOLAR":
                old_station_name="KSPDCL"
                meter_list=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Station_Name','Description','Meter_No')
                meter_list=meter_list.exclude(Description__icontains="STANDBY").values_list('Station_Name','Description','Meter_No')
            elif old_station_name =="AZURE" :
                meter_list=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Meter_No')
                meter_list=meter_list.exclude(Description__icontains="ANANTHAPURAM").values_list('Meter_No')
            elif old_station_name =="TATA":
                meter_list=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Meter_No')
                meter_list=meter_list.exclude(Description__icontains="ANANTHAPURAM").values_list('Meter_No')
            elif old_station_name =="FORTUM":
                meter_list=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Meter_No')
                meter_list=meter_list.exclude(Description__icontains="FINN").values_list('Meter_No')
            
            elif old_station_name in pvg_solarlist:
                meter_list=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Meter_No')
            else:
                        
                total_details= list(StationDetails.objects.filter(station_name__in=[old_station_name],dateofchecking__range=(start_date, end_date),meter_difference__range=[new_threshold,'23:59:59']).values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference','dateofchecking','correction_needed','meter_status').order_by('-meter_difference').distinct())

                location_set=[]
                for i in range(0,len(total_details)):
                    location=[]
                    location=Semlog_data.objects.filter(Description__icontains=total_details[i][2]).values_list('Location','Description')
                   
                    location_set.append(location[0])
                
                for i in range(0,len(total_details)):
                    stat=list(total_details[i])
                    stat[4]=str(stat[4])
                    stat[5]=str(stat[5])
                    stat[6]=str(stat[6])
                    stat.insert(4,str(location_set[i][0]))    
                    total_details[i]=tuple(stat)
                
                return render(request5, 'meter_details.html', {'totdetails':total_details})

                # meter_list=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Meter_No')

            modified_list=[]
            for i in range(0,len(meter_list)):
                modified_list.append(meter_list[i][0])


            total_details=list(StationDetails.objects.filter(Meter_no__in=[i for i in modified_list],dateofchecking__range=(start_date, end_date),meter_difference__range=[new_threshold,'23:59:59']).values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference','dateofchecking','correction_needed','meter_status').order_by('-meter_difference').distinct())
            location_set=[]
            for i in range(0,len(total_details)):
                location=[]
                location=Semlog_data.objects.filter(Description__icontains=total_details[i][2]).values_list('Location','Description')
                   
                location_set.append(location[0])
           
            for i in range(0,len(total_details)):
                stat=list(total_details[i])
                stat[4]=str(stat[4])
                stat[5]=str(stat[5])
                stat[6]=str(stat[6])
                stat.insert(4,str(location_set[i][0]))    
                total_details[i]=tuple(stat)
                
            return render(request5, 'meter_details.html', {'totdetails': total_details})

    except:
        raise Http404("Page does not exist")
@login_required()
def date_range1(request1):
    try:

        if request1.method=="POST":
            start_date=str(request1.POST.get('fromdate')),
            end_date=str(request1.POST.get('todate')),
           
            if request1.user.username == "SRLDC_MO":
                threshold=str(request1.POST.get('threshold')),

            else:
                threshold =["00:00:00",]
           
            new_start_date=datetime.datetime.strptime(start_date[0],'%m/%d/%Y').date(),
            new_end_date=datetime.datetime.strptime(end_date[0],'%m/%d/%Y').date(),
           
            new_threshold=datetime.datetime.strptime(threshold[0],'%H:%M:%S').time()
            
            request1.session['start_dat']=start_date
            request1.session['end_dat']=end_date
            request1.session['excel_threshold']=threshold

            global sr1_list,sr2_list,pvg_solar,pvg_solarlist,station_names

            old_station_name=request1.user.username
            

            
            if request1.user.username =='SRLDC_MO':
                date_wise_details=list(StationDetails.objects.filter(dateofchecking__range=(new_start_date[0], new_end_date[0]),meter_difference__range=(new_threshold,'23:59:59')).values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference','dateofchecking','correction_needed','meter_status').order_by('-meter_difference').distinct())

                location_set=[]
                for i in range(0,len(date_wise_details)):
                    location=[]
                    location=Semlog_data.objects.filter(Description__icontains=date_wise_details[i][2]).values_list('Location','Description')
                   
                    location_set.append(location[0])


                
                for i in range(0,len(date_wise_details)):
                    stat=list(date_wise_details[i])
                    stat[4]=str(stat[4])
                    stat[5]=str(stat[5])
                    stat[6]=str(stat[6])
                    
                    stat.insert(4,str(location_set[i][0]))
                    date_wise_details[i]=tuple(stat)
            
                # gsggfd
                
                return render(request1, 'date_range.html', {'datedetails':date_wise_details,'start_date1':start_date[0],'end_date1':end_date[0]})
            
            elif request1.user.username =='S_SR1PG':

                date_wise_details= list(StationDetails.objects.filter(station_name__in=[i for i in sr1_list],dateofchecking__range=(new_start_date[0], new_end_date[0]),meter_difference__range=(new_threshold,'23:59:59')).values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference','dateofchecking','correction_needed','meter_status').order_by('-meter_difference').distinct())
                for i in range(0,len(date_wise_details)):
                    stat=list(date_wise_details[i])
                    stat[4]=str(stat[4])
                    stat[5]=str(stat[5])
                    stat[6]=str(stat[6])
                        
                    date_wise_details[i]=tuple(stat)
                
                return render(request1, 'date_range.html', {'datedetails':date_wise_details,'start_date1':start_date[0],'end_date1':end_date[0]})

            elif request1.user.username =='S_SR2PG':
                date_wise_details= list(StationDetails.objects.filter(station_name__in=[i for i in sr2_list],dateofchecking__range=(new_start_date[0], new_end_date[0]),meter_difference__range=(new_threshold,'23:59:59')).values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference','dateofchecking','correction_needed','meter_status').order_by('-meter_difference').distinct())
                
                for i in range(0,len(date_wise_details)):
                    stat=list(date_wise_details[i])
                    stat[4]=str(stat[4])
                    stat[5]=str(stat[5])
                    stat[6]=str(stat[6])
                        
                    date_wise_details[i]=tuple(stat)
                
                return render(request1, 'date_range.html', {'datedetails':date_wise_details,'start_date1':start_date[0],'end_date1':end_date[0]})


            else:
                global station_names,pvg_solar,pvg_solarlist
                if old_station_name not in station_names:
                    if old_station_name not in pvg_solar:

                        old_station_name=old_station_name
                    else: 
                        old_station_name=pvg_solar[old_station_name]
                else:
                    old_station_name=station_names[old_station_name]

                if old_station_name =="KSPDCL":
                    meter_list=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Meter_No')
                    meter_list=meter_list.exclude(Description__icontains="MAIN").values_list('Meter_No')
                    meter_list=meter_list.exclude(Description__icontains="CHECK").values_list('Meter_No')
                elif old_station_name =="PG_SOLAR":
                    old_station_name="KSPDCL"
                    meter_list=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Station_Name','Description','Meter_No')
                    meter_list=meter_list.exclude(Description__icontains="STANDBY").values_list('Station_Name','Description','Meter_No')
                elif old_station_name =="TATA":
                    meter_list=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Meter_No')
                    meter_list=meter_list.exclude(Description__icontains="ANANTHAPURAM").values_list('Meter_No')
                elif old_station_name =="AZURE" :
                    meter_list=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Meter_No')
                    meter_list=meter_list.exclude(Description__icontains="ANANTHAPURAM").values_list('Meter_No')
                elif old_station_name =="FORTUM":
                    meter_list=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Meter_No')
                    meter_list=meter_list.exclude(Description__icontains="FINN").values_list('Meter_No')
                elif old_station_name in pvg_solarlist:
                    meter_list=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Meter_No')

    
                else:
                    date_wise_details= list(StationDetails.objects.filter(station_name__in=[old_station_name],dateofchecking__range=(new_start_date[0], new_end_date[0]),meter_difference__range=(new_threshold,'23:59:59')).values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference','dateofchecking','correction_needed','meter_status').order_by('-meter_difference').distinct())

                    location_set=[]
                    for i in range(0,len(date_wise_details)):
                        location=[]
                        location=Semlog_data.objects.filter(Description__icontains=date_wise_details[i][2]).values_list('Location','Description')
                   
                        location_set.append(location[0])

                
                    for i in range(0,len(date_wise_details)):
                        stat=list(date_wise_details[i])
                        stat[4]=str(stat[4])
                        stat[5]=str(stat[5])
                        stat[6]=str(stat[6])
                        stat.insert(4,str(location_set[i][0]))
                        
                        date_wise_details[i]=tuple(stat)
                
                    return render(request1, 'date_range.html', {'datedetails':date_wise_details,'start_date1':start_date[0],'end_date1':end_date[0]})
                    # meter_list=Semlog_data.objects.filter(Description__icontains=old_station_name).values_list('Meter_No')

                modified_list=[]
                for i in range(0,len(meter_list)):
                    modified_list.append(meter_list[i][0])


                date_wise_details=list(StationDetails.objects.filter(dateofchecking__range=(new_start_date[0], new_end_date[0]),meter_difference__range=(new_threshold,'23:59:59'),
                Meter_no__in =[i for i in modified_list]).values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference','dateofchecking','correction_needed','meter_status').order_by('-meter_difference').distinct())

                location_set=[]
                for i in range(0,len(date_wise_details)):
                    location=[]
                    location=Semlog_data.objects.filter(Description__icontains=date_wise_details[i][2]).values_list('Location','Description')
                   
                    location_set.append(location[0])

                for i in range(0,len(date_wise_details)):
                    stat=list(date_wise_details[i])
                    stat[4]=str(stat[4])
                    stat[5]=str(stat[5])
                    stat[6]=str(stat[6])

                    stat.insert(4,str(location_set[i][0]))
                        
                    date_wise_details[i]=tuple(stat)

    
                return render(request1,'date_range.html', {'datedetails':date_wise_details,'start_date1':start_date[0],'end_date1':end_date[0]})



        else:
            return HttpResponseNotFound(request1,'error_404.html')
    except:
        raise Http404("Page does not exist")
@login_required()
def drift_range(request1):
    if request1.method=='POST':
        start_date=str(request1.POST.get('fromdate')),
        end_date=str(request1.POST.get('todate')),
        new_start_date=datetime.datetime.strptime(start_date[0],'%m/%d/%Y').date(),
        new_end_date=datetime.datetime.strptime(end_date[0],'%m/%d/%Y').date(),

        drift_speed=request1.POST.get('driftvalue'),
        drift_speed=drift_speed[0]
        drift_speed1=datetime.datetime.strptime(drift_speed,'%H:%M:%S').time()
        
        request1.session['new_date']=start_date[0]
        request1.session['end_date']=end_date[0]
        request1.session['drift_speed']=drift_speed

        date_wise_details=StationDetails.objects.filter(dateofchecking__range=(new_start_date[0], new_end_date[0])).values_list('utility_name','station_name','Meter_no','gps','meter_drift','meter_difference','dateofchecking','correction_needed','meter_status').values()
                
                
        date_wise_details=list(date_wise_details.filter(meter_difference__range=[drift_speed1,'23:59:59']).values_list('utility_name','station_name','Meter_no','gps','meter_drift','meter_difference','dateofchecking','correction_needed','meter_status'))

               
                
        for i in range(0,len(date_wise_details)):
            stat=list(date_wise_details[i])
            stat[3]=str(stat[3])
            stat[4]=str(stat[4])
            stat[5]=str(stat[5])
                    
            date_wise_details[i]=tuple(stat)
            
        return render(request1, 'drift_details.html', {'driftdetails':date_wise_details})
            
    


    else:
        return render(request1,'failure.html')
@login_required()
def drift_excel_download(request1):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Timedrift_range.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Td_Range')
    # Styling of sheet
    
    New_date=request1.session['new_date']
    End_date=request1.session['end_date']
    New_driftspeed=request1.session['drift_speed']
    
    new_start_date=datetime.datetime.strptime(New_date,'%m/%d/%Y').date(),
    new_end_date=datetime.datetime.strptime(End_date,'%m/%d/%Y').date(),

    
    drift_speed1=datetime.datetime.strptime(New_driftspeed,'%H:%M:%S').time()
    
    def set_style(name, height, bold=True, center=True):
        style = xlwt.XFStyle()  # ?????

        font = xlwt.Font()  # ???????
        font.name = name  # 'Times New Roman'
        font.bold = bold
        font.color_index = 0
        font.height = height

        borders= xlwt.Borders()
        borders.left= 6
        borders.right= 6
        borders.top= 6
        borders.bottom= 6

        style.font = font
    # style.borders = borders
        alignment = xlwt.Alignment()
        alignment.horz = xlwt.Alignment.HORZ_CENTER
        if center == True:
            style.alignment = alignment
        return style 



    # Sheet header, first row
    row_num = 0
    columns = ['Utility_Name','Station_Name','Meter_No', 'GPS', 'Meter_time','Time Drift', 'Date of Check(y-m-d)', 'Correction Needed','Meter_Status(Fast/Slow)']

    for col_num in range(len(columns)):
        ws.col(col_num).width = 6500
        ws.write(row_num, col_num, columns[col_num],set_style('Times New Roman', 280, True, True))

    # Sheet body, remaining rows
    default_style = xlwt.XFStyle()

    # new body
    rows=StationDetails.objects.filter(dateofchecking__range=(new_start_date[0], new_end_date[0]),meter_difference__range=[drift_speed1,'23:59:59']).values_list('utility_name','station_name','Meter_no',
                                                'gps',
                                                'meter_drift','meter_difference',
                                                'dateofchecking',
        
                                                'correction_needed','meter_status')
    
    rows= rows.filter().order_by('-meter_difference').values_list('utility_name','station_name','Meter_no',
                                                'gps',
                                                'meter_drift','meter_difference',
                                                'dateofchecking',
        
                                                'correction_needed','meter_status')

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), set_style('Times New Roman', 250, False, True))
    
    wb.save(response)
    return response

@login_required()
def station_fetch(request1):
    try:

        if request1.user.username=="SRLDC_MO":
            
            updated_stations =Station_store.objects.filter().order_by('station_name').values_list('station_name')
            updated_list=[]
            for i in range(len(updated_stations)):
                updated_list.append(updated_stations[i][0])

            return render(request1, 'index1.html', {'updated_details': updated_list})
        else:
            return render(request1,'error_404.html')
    except:
        raise Http404("Page does not exist")
@login_required()
def add_station(request):
    
    return render(request,'add_station.html')  
@login_required()
def adding_station(request):
    if request.method=='GET':
        utility_name=request.GET.get('utilityname')
        station_name=request.GET.get('stationname')
        utility_name=utility_name.upper()
        station_name=station_name.upper()
        stat_data=Station_store(
                        utility_name=utility_name,
                        station_name=station_name
        )
        stat_data.save()
        return render(request,'success.html') 
    else:
        return render(request,'error_404.html') 
@login_required()
def download_excel_data(request1):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Timedrift.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Meter_details')
    # Styling of sheet
    global sr1_list,sr2_list
    def set_style(name, height, bold=True, center=True):
        style = xlwt.XFStyle()  # ?????

        font = xlwt.Font()  # ???????
        font.name = name  # 'Times New Roman'
        font.bold = bold
        font.color_index = 0
        font.height = height

        borders= xlwt.Borders()
        borders.left= 6
        borders.right= 6
        borders.top= 6
        borders.bottom= 6

        style.font = font
    # style.borders = borders
        alignment = xlwt.Alignment()
        alignment.horz = xlwt.Alignment.HORZ_CENTER
        if center == True:
            style.alignment = alignment
        return style 



    # Sheet header, first row
    row_num = 0
    columns = ['Station_Name','Meter_No', 'GPS', 'Meter_time', 'DOC(YYYY-MM-DD)', 'Correction Needed','Remarks']

    for col_num in range(len(columns)):
        ws.col(col_num).width = 6500
        ws.write(row_num, col_num, columns[col_num],set_style('Times New Roman', 280, True, True))

    # Sheet body, remaining rows
    default_style = xlwt.XFStyle()
    station_name=request1.user.username
    if request1.user.username =='SRLDC_MO':
    

        rows=StationDetails.objects.all().values_list('station_name',
                                                'gps',
                                                'meter_drift',
                                                str('dateofchecking'),
        
                                                'correction_needed')
        rows= rows.filter().order_by('-meter_drift').values_list('station_name',
                                                'gps',
                                                'meter_drift',
                                                str('dateofchecking'),
        
                                                'correction_needed')

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, str(row[col_num]), set_style('Times New Roman', 250, False, True))

        
    elif request1.user.username =='S_SR1PG':
    

        rows=StationDetails.objects.filter(station_name__in=[i for i in sr1_list]).values_list('station_name',
                                                'gps',
                                                'meter_drift',
                                                str('dateofchecking'),
        
                                                'correction_needed')
        rows =rows.filter().order_by('-meter_drift').values_list('station_name',
                                                'gps',
                                                'meter_drift',
                                                str('dateofchecking'),
        
                                                'correction_needed')
        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, str(row[col_num]), set_style('Times New Roman', 250, False, True))

        
    elif request1.user.username =='S_SR2PG':
    

        rows=StationDetails.objects.filter(station_name__in=[i for i in sr2_list]).values_list('station_name',
                                                'gps',
                                                'meter_drift',
                                                str('dateofchecking'),
        
                                                'correction_needed')
        rows =rows.filter().order_by('-meter_drift').values_list('station_name',
                                                'gps',
                                                'meter_drift',
                                                str('dateofchecking'),
        
                                                'correction_needed')

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, str(row[col_num]), set_style('Times New Roman', 250, False, True))

        
    else:
        rows=StationDetails.objects.filter(station_name__in=[station_name]).values_list('station_name','Meter_no',
                                                'gps',
                                                'meter_drift',
                                                str('dateofchecking'),
        
                                                'correction_needed','remarks')

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, str(row[col_num]), set_style('Times New Roman', 250, False, True))

        
    wb.save(response)
    return response
@login_required()     
def date_wise_excel(request1):
    response = HttpResponse(content_type='application/ms-excel')
    
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Date_wise_details')
    start_date=request1.session['start_dat']
    end_date=request1.session['end_dat']
    new_threshold=request1.session['excel_threshold']
    new_start_date=datetime.datetime.strptime(start_date[0],'%m/%d/%Y').date(),
    new_end_date=datetime.datetime.strptime(end_date[0],'%m/%d/%Y').date(),
    new_threshold=datetime.datetime.strptime(new_threshold[0],'%H:%M:%S').time(),
    response['Content-Disposition'] = 'attachment; filename="Tdrift_date_wise.xls"'
    
    global sr1_list,sr2_list,station_names,pvg_solar,pvg_solarlist
    sub_station=str(request1.user.username)
    # Sheet header, first row
    row_num = 0
    # Styling to sheet
    def set_style(name, height, bold=False, center=True):
        style = xlwt.XFStyle()  # ?????

        font = xlwt.Font()  # ???????
        font.name = name  # 'Times New Roman'
        font.bold = bold
        font.color_index = 0
        font.height = height

        borders= xlwt.Borders()
        borders.left= 6
        borders.right= 6
        borders.top= 6
        borders.bottom= 6

        style.font = font
    # style.borders = borders
        alignment = xlwt.Alignment()
        #alignment.horz = xlwt.Alignment.HORZ_CENTER
        if center == True:
            style.alignment = alignment
        return style 


    columns = ['Utility_Name','Station_Name','Description','Meter_No', 'GPS', 'Meter Time','Time Drift', 'Date Of Checking', 'Correction done or not','Meter-status(Fast/Slow)']

    for col_num in range(len(columns)):
        ws.col(col_num).width = 6500
        ws.write(row_num, col_num, columns[col_num], set_style('Times New Roman', 250, True, True))

    # Sheet body, remaining rows
    default_style = xlwt.XFStyle()
    station_name=request1.user.username
    

    if request1.user.username =='SRLDC_MO':
        rows=StationDetails.objects.filter(dateofchecking__range=(new_start_date[0], new_end_date[0]),meter_difference__range=(new_threshold[0],'23:59:59')).values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference',str('dateofchecking'),'correction_needed','meter_status').order_by('-meter_difference').distinct()
        
        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, str(row[col_num]), set_style('Times New Roman', 250, False, True))

    elif request1.user.username =='S_SR1PG':
        rows=StationDetails.objects.filter(station_name__in=[i for i in sr1_list],dateofchecking__range=(new_start_date[0], new_end_date[0]),meter_difference__range=(new_threshold[0],'23:59:59')).values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference',str('dateofchecking'),'correction_needed','meter_status').order_by('-meter_difference').distinct()

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, str(row[col_num]), set_style('Times New Roman', 250, False, True))

    elif request1.user.username =='S_SR2PG':
        rows=StationDetails.objects.filter(station_name__in=[i for i in sr2_list],dateofchecking__range=(new_start_date[0], new_end_date[0]),meter_difference__range=(new_threshold[0],'23:59:59')).values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference',str('dateofchecking'),'correction_needed','meter_status').order_by('-meter_difference').distinct()

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, str(row[col_num]), set_style('Times New Roman', 250, False, True))
    else:

        if station_name not in station_names:
            if station_name not in pvg_solar:
                station_name=station_name
            else:
                station_name=pvg_solar[station_name]
        else:
            station_name=station_names[station_name]



        if station_name =="KSPDCL":
            meter_list=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Meter_No')
            meter_list=meter_list.exclude(Description__icontains="MAIN").values_list('Meter_No')
            meter_list=meter_list.exclude(Description__icontains="CHECK").values_list('Meter_No')
        elif station_name =="PG_SOLAR":
            station_name="KSPDCL"
            meter_list=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Station_Name','Description','Meter_No')
            meter_list=meter_list.exclude(Description__icontains="STANDBY").values_list('Station_Name','Description','Meter_No')
        elif station_name =="AZURE" :
            meter_list=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Meter_No')
            meter_list=meter_list.exclude(Description__icontains="ANANTHAPURAM").values_list('Meter_No')
        elif station_name =="TATA":
            meter_list=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Meter_No')
            meter_list=meter_list.exclude(Description__icontains="ANANTHAPURAM").values_list('Meter_No')
        elif request1.user.username =="FORTUM":
            meter_list=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Meter_No')
            meter_list=meter_list.exclude(Description__icontains="FINN").values_list('Meter_No')
        elif station_name in pvg_solarlist:
            meter_list=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Meter_No')
        else:
            meter_list=Semlog_data.objects.filter(Station_Name__icontains=station_name).values_list('Meter_No')

            

        modified_list=[]
        for i in range(0,len(meter_list)):
            modified_list.append(meter_list[i][0])

        rows=StationDetails.objects.filter(Meter_no__in=[i for i in modified_list],dateofchecking__range=(new_start_date[0], new_end_date[0]),meter_difference__range=(new_threshold[0],'23:59:59')).values_list('utility_name','station_name','description','Meter_no','gps','meter_drift','meter_difference',str('dateofchecking'),'correction_needed','meter_status').order_by('-meter_difference').distinct()

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, str(row[col_num]), set_style('Times New Roman', 250, False, True))

    wb.save(response)
    return response 
@login_required()
def file_format(request):
    station_name=request.session['station1']
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Upload_file.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Sem_details')
    # Styling of sheet
    
    def set_style(name, height, bold=True, center=True):
        style = xlwt.XFStyle()  # ?????

        font = xlwt.Font()  # ???????
        font.name = name  # 'Times New Roman'
        font.bold = bold
        font.color_index = 0
        font.height = height

        borders= xlwt.Borders()
        borders.left= 6
        borders.right= 6
        borders.top= 6
        borders.bottom= 6

        style.font = font
    # style.borders = borders
        alignment = xlwt.Alignment()
        alignment.horz = xlwt.Alignment.HORZ_CENTER
        if center == True:
            style.alignment = alignment
        return style 



    # Sheet header, first row
    row_num = 0
    columns = ['Station_name','Description','Meter_no','GPS(hh:mm:ss)', 'Meter_time(hh:mm:ss)', 'Date_of_Check(dd/mm/yy)', 'Correction_Needed(Yes/No)','Remarks']

    for col_num in range(len(columns)):
        ws.col(col_num).width = 8500
        ws.write(row_num, col_num, columns[col_num],set_style('Times New Roman', 280, True, True))

    # Sheet body, remaining rows
    default_style = xlwt.XFStyle()
    global station_names,pvg_solar,pvg_solarlist

    if station_name not in station_names:
        if station_name not in pvg_solar:
            station_name=station_name
        else:
            station_name=pvg_solar[station_name]
    else:
        station_name=station_names[station_name]

    
    # if request.user.username =='SRLDC_MO':
    if station_name =="KSPDCL":
        new=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Station_Name','Description','Meter_No')
        new=new.exclude(Description__icontains="MAIN").values_list('Station_Name','Description','Meter_No')
        new=new.exclude(Description__icontains="CHECK").values_list('Station_Name','Description','Meter_No')
    elif station_name=="PG_SOLAR":
        station_name="KSPDCL"
        new=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Station_Name','Description','Meter_No')
        new=new.exclude(Description__icontains="STANDBY").values_list('Station_Name','Description','Meter_No')

    elif station_name=="AZURE":
        new=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Station_Name','Description','Meter_No')
        new=new.exclude(Description__icontains="ANANTHAPURAM").values_list('Station_Name','Description','Meter_No')
    
    elif station_name =="FORTUM":
        new=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Station_Name','Description','Meter_No')
        new=new.exclude(Description__icontains="FINN").values_list('Station_Name','Description','Meter_No')
    elif station_name=="TATA":
        new=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Station_Name','Description','Meter_No')
        new=new.exclude(Description__icontains="ANANTHAPURAM").values_list('Station_Name','Description','Meter_No')
    elif station_name in pvg_solarlist:
        new=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Station_Name','Description','Meter_No')

    

        # Ananthapuram solar stations
    elif station_name=="ACMEKRNL":
        new=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Station_Name','Description','Meter_No')
    elif station_name=="ACMEHIS":
        new=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Station_Name','Description','Meter_No')
    elif station_name=="ACMEBHI":
        new=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Station_Name','Description','Meter_No')
    elif station_name=="FRV":
        new=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Station_Name','Description','Meter_No')
        new=new.exclude(Description__icontains="FRV8").values_list('Station_Name','Description','Meter_No') 
    elif station_name=="FRV8":
        new=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Station_Name','Description','Meter_No')   

    elif station_name=="TATA_ANTP":
        station_name="TATA"
        new=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Station_Name','Description','Meter_No')
        new=new.exclude(Description__icontains="PAVAGADA").values_list('Station_Name','Description','Meter_No')
    elif station_name=="AZURE_ANTP":
        station_name="AZURE"
        new=Semlog_data.objects.filter(Description__icontains=station_name).values_list('Station_Name','Description','Meter_No')
        new=new.exclude(Description__icontains="PAVAGADA").values_list('Station_Name','Description','Meter_No')         
    else:

        new=Semlog_data.objects.filter(Station_Name__in=[station_name]).values_list('Station_Name','Description','Meter_No')

    for row in new:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), set_style('Times New Roman', 250, False, True))
    
    wb.save(response)
    return response 
@login_required()
def final_status(request):
    try:
        
        state_entity=request.GET.get('state')
        month1=request.GET.get('month')
        year=int(month1[3:])
        
        first_day=[]
        last_day=[]
        
        request.session["selected_month"]=month1
        global sr1_list,sr2_list,solar
        def last_day_of_month(any_day):
            next_month = any_day.replace(day=28) + datetime.timedelta(days=4)
            return next_month - datetime.timedelta(days=next_month.day)
        
        for month in range(1, 13):
            todayDate = datetime.date(year,month,1)
            if todayDate.day > 25:
                todayDate += datetime.timedelta(7)
            first_day.append(todayDate.replace(day=1))
            last_day.append(last_day_of_month(datetime.date(year, month, 1)))
        if month1[0:2]=='01':
            start_date=first_day[0]
            end_date=last_day[0]
        elif month1[0:2]=='02':
            start_date=first_day[1]
            end_date=last_day[1]
        elif month1[0:2]=='03':
            start_date=first_day[2]
            end_date=last_day[2]
        elif month1[0:2]=='04':
            start_date=first_day[3]
            end_date=last_day[3]
        elif month1[0:2]=='05':
            start_date=first_day[4]
            end_date=last_day[4]
        elif month1[0:2]=='06':
            start_date=first_day[5]
            end_date=last_day[5]
        elif month1[0:2]=='07':
            start_date=first_day[6]
            end_date=last_day[6]
        elif month1[0:2]=='08':
            start_date=first_day[7]
            end_date=last_day[7]
        elif month1[0:2]=='09':
            start_date=first_day[8]
            end_date=last_day[8]
        elif month1[0:2]=='10':
            start_date=first_day[9]
            end_date=last_day[9]
        elif month1[0:2]=='11':
            start_date=first_day[10]
            end_date=last_day[10]
        else:
            start_date=first_day[11]
            end_date=last_day[11]
        # old_station_name=request5.session['station1']
        station_name=request.session['admin']
        
        if request.user.username =='SRLDC_MO':
    
            # Station_details=StationDetails.objects.values_list('station_name','dateofchecking').values()
            # Station_details1=list(StationDetails.objects.filter(description__icontains=["KREDL","ADANI"]).values_list('utility_name','station_name','dateofupload'))
            
            Station_details=list(StationDetails.objects.filter(utility_name__in=[state_entity],dateofupload__range=(start_date, end_date)).order_by('station_name').values_list('utility_name','station_name','dateofupload'))

            Final_list=[]
            for i in range(0,13):
                Station_list=[]
                station_name=pvg_solarlist[i]
                Station_details1=StationDetails.objects.filter(description__icontains=station_name,dateofupload__range=(start_date, end_date)).order_by('dateofupload').values_list('utility_name','description','dateofupload').distinct()
                if len(Station_details1) > 0:
                    Final_list.append((Station_details1[0][0],station_name,Station_details1[len(Station_details1)-1][2],"YES"))
                    
                else:
                    
                    Final_list.append(("SOLAR",station_name,"NULL","NO"))
                
            if state_entity =="SOLAR":
                return render(request, 'monthlyfile_status.html', {'totdetails': Final_list})

            else:

                Station_list=list(Semlog_data.objects.filter(Utility_Name__in=[state_entity]).values_list('Utility_Name','Station_Name').distinct())
                
                if len(Station_details)!= 0:
                    
                    Station_list1=[]
                    for i in range(0,len(Station_details)):
                        Station_list1.append(Station_details[i][1])
                    
                    Station_list2=[]
                    
                    for i in range(0,len(Station_list)):
                        Station_list2.append((Station_list[i][0],Station_list[i][1]))
            
                    
                    stat=[]
                    
                    for i in range(0,len(Station_details)):
                        stat=list(Station_details[i])
                        stat.append("YES")
                        # stat.append(Station_details[i][1])
                        Station_details[i]=tuple(stat)
                    
                    
                    
                    for i in range(0,len(Station_list2)):

                        if Station_list2[i][1] not in Station_list1:
                            Station_details.append((Station_list2[i][0],Station_list2[i][1],'NULL','NO'))
                        else:
                            continue
                    
                else:
                    for i in range(0,len(Station_list)):

                        Station_details.append((Station_list[i][0],Station_list[i][1],'NULL','NO'))
                        

                Station_details=list(set([i for i in Station_details]))
                Station_details=sorted(Station_details, key=lambda x: x[0]) 
                    
                return render(request, 'monthlyfile_status.html', {'totdetails': Station_details})
                
        elif request.user.username =='S_SR1PG':
            Station_details=list(StationDetails.objects.filter(station_name__in=[i for i in sr1_list],dateofchecking__range=(start_date, end_date),).order_by('-dateofchecking').values_list('utility_name','station_name','dateofupload'))
            
            

            Station_list=list(Semlog_data.objects.filter(Station_Name__in=[i for i in sr1_list]).values_list('Utility_Name','Station_Name').distinct())
            
            if len(Station_details)!=0:

                Station_list1=[]
                for i in range(0,len(Station_details)):
                    Station_list1.append(Station_details[i][0])

                Station_list2=[]
                for i in range(0,len(Station_list)):
                    Station_list2.append((Station_list[i][0],Station_list[i][1]))
                

                for i in range(0,len(Station_details)):
                    # if Station_list2[i] in Station_list1:
                        stat=list(Station_details[i])
                        stat.append("YES")
                        Station_details[i]=tuple(stat)
                
                for i in range(0,len(Station_list2)):

                    if Station_list2[i] not in Station_list1:
                        Station_details.append((Station_list2[i][0],Station_list2[i][1],'NULL','NO'))
                    else:
                        continue
            else:
                for i in range(0,len(Station_list)):

                    Station_details.append((Station_list[i][0],Station_list[i][1],'NULL','NO'))

            
            Station_details=list(set([i for i in Station_details]))
            Station_details=sorted(Station_details, key=lambda x: x[0])
            
            return render(request, 'powergrid_status.html', {'totdetails': Station_details})
        
        
        
        elif request.user.username =='S_SR2PG':
            
            Station_details=list(StationDetails.objects.filter(station_name__in=[i for i in sr2_list],dateofchecking__range=(start_date, end_date)).order_by('-dateofchecking').values_list('utility_name','station_name','dateofupload'))


            Station_list=list(Semlog_data.objects.filter(Station_Name__in=[i for i in sr2_list]).values_list('Utility_Name','Station_Name').distinct())
            
            if len(Station_details)!=0:
                Station_list1=[]
                for i in range(0,len(Station_details)):
                    Station_list1.append(Station_details[i][0])

                Station_list2=[]
                for i in range(0,len(Station_list)):
                    Station_list2.append((Station_list[i][0],Station_list[i][1]))
                

                for i in range(0,len(Station_details)):
                    # if Station_list2[i] in Station_list1:
                        stat=list(Station_details[i])
                        stat.append("YES")
                        Station_details[i]=tuple(stat)
                
                for i in range(0,len(Station_list2)):

                    if Station_list2[i] not in Station_list1:
                        Station_details.append((Station_list2[i][0],Station_list2[i][1],'NULL','NO'))
                    else:
                        continue
            else:
                for i in range(0,len(Station_list)):

                    Station_details.append((Station_list[i][0],Station_list[i][1],'NULL','NO'))
                        
            Station_details=list(set([i for i in Station_details]))
            Station_details=sorted(Station_details, key=lambda x: x[0])
            return render(request, 'powergrid_status.html', {'totdetails': Station_details})

    # For individual station stauts ,now it is not needed.when it rquires just add below code


        # else:
        #     Station_details=list(StationDetails.objects.filter(station_name__in=[station_name],dateofchecking__range=(start_date, end_date)).order_by('-dateofchecking').values_list('utility_name','station_name','dateofupload'))
        #     Station_list=list(Semlog_data.objects.filter().values_list('Utility_Name','Station_Name').distinct())
        #     Station_list1=[]
        #     for i in range(0,len(Station_details)):
        #         Station_list1.append(Station_details[i][0])

        #     Station_list2=[]
        #     for i in range(0,len(Station_list)):
        #         Station_list2.append((Station_list[i][0],Station_list[i][1]))
            
        #     for i in range(0,len(Station_details)):
        #         # if Station_list2[i] in Station_list1:
        #             stat=list(Station_details[i])
        #             stat.append("YES")
        #             Station_details[i]=tuple(stat)
            
        #     for i in range(0,len(Station_list2)):

        #         if Station_list2[i] not in Station_list1:
        #             Station_details.append((Station_list2[i][0],Station_list2[i][1],'NULL','NO'))
        #         else:
        #             continue
                    
        #     Station_details=list(set([i for i in Station_details]))
        #     Station_details=sorted(Station_details, key=lambda x: x[0])
        #     return render(request, 'month_filelist.html', {'totdetails': Station_details})
    except:
        raise Http404("Page does not exist")
